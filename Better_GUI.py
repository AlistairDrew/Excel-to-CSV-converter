import openpyxl
import tkinter as tk
from tkinter import filedialog

def get_column_letter(column_index):
    """Get the column letter from the column index"""
    dividend = column_index
    column_name = ''
    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_name = chr(65 + modulo) + column_name
        dividend = (dividend - modulo) // 26
    return column_name

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

headings_window = tk.Toplevel()
headings_window.title('Manage Column Headings')

current_headings_label = tk.Label(headings_window, text='Current Column Headings:')
current_headings_label.grid(row=0, column=0)

current_headings_listbox = tk.Listbox(headings_window)
for col in worksheet.iter_cols(1, worksheet.max_column):
    current_headings_listbox.insert(tk.END, col[0].value)
current_headings_listbox.grid(row=1, column=0)

new_heading_label = tk.Label(headings_window, text='Add New Column Heading:')
new_heading_label.grid(row=2, column=0)

new_heading_entry = tk.Entry(headings_window)
new_heading_entry.grid(row=3, column=0)

def add_heading():
    new_heading = new_heading_entry.get()
    if new_heading:
        column_index = worksheet.max_column + 1
        worksheet.cell(row=1, column=column_index, value=new_heading)
        current_headings_listbox.insert(tk.END, new_heading)
        new_heading_entry.delete(0, tk.END)

add_heading_button = tk.Button(headings_window, text='Add', command=add_heading)
add_heading_button.grid(row=3, column=1)

def import_headings():
    import_file_path = filedialog.askopenfilename()
    import_workbook = openpyxl.load_workbook(import_file_path)
    import_worksheet = import_workbook.active
    for col in import_worksheet.iter_cols(1, import_worksheet.max_column):
        worksheet.cell(row=1, column=worksheet.max_column+1, value=col[0].value)
        current_headings_listbox.insert(tk.END, col[0].value)

import_button = tk.Button(headings_window, text='Import Headings', command=import_headings)
import_button.grid(row=4, column=0)

def clear_headings():
    for col in worksheet.iter_cols(1, worksheet.max_column):
        col[0].value = None
    current_headings_listbox.delete(0, tk.END)

clear_button = tk.Button(headings_window, text='Clear', command=clear_headings)
clear_button.grid(row=4,



