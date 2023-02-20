import openpyxl
import tkinter as tk
from tkinter import filedialog

# Define a Tkinter window for the file dialog
root = tk.Tk()
root.withdraw()

# Ask the user to select an excel file
file_path = filedialog.askopenfilename()

# Load the selected excel file
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

# Define a Tkinter window for managing the column headings
headings_window = tk.Toplevel()
headings_window.title('Manage Column Headings')

# Create a label for the current column headings
current_headings_label = tk.Label(headings_window, text='Current Column Headings:')
current_headings_label.grid(row=0, column=0)

# Create a listbox for displaying the current column headings
current_headings_listbox = tk.Listbox(headings_window)
for col in worksheet.iter_cols(1, worksheet.max_column):
    current_headings_listbox.insert(tk.END, col[0].value)
current_headings_listbox.grid(row=1, column=0)

# Create a label and entry for adding a new column heading
new_heading_label = tk.Label(headings_window, text='Add New Column Heading:')
new_heading_label.grid(row=2, column=0)
new_heading_entry = tk.Entry(headings_window)
new_heading_entry.grid(row=3, column=0)

# Create a button for adding a new column heading
def add_heading():
    new_heading = new_heading_entry.get()
    if new_heading:
        column_index = worksheet.max_column + 1
        worksheet.cell(row=1, column=column_index, value=new_heading)
        current_headings_listbox.insert(tk.END, new_heading)
        new_heading_entry.delete(0, tk.END)

add_heading_button = tk.Button(headings_window, text='Add', command=add_heading)
add_heading_button.grid(row=3, column=1)

# Create a button for importing column headings
def import_headings():
    import_file_path = filedialog.askopenfilename()
    import_workbook = openpyxl.load_workbook(import_file_path)
    import_worksheet = import_workbook.active
    for col in import_worksheet.iter_cols(1, import_worksheet.max_column):
        worksheet.cell(row=1, column=worksheet.max_column+1, value=col[0].value)
        current_headings_listbox.insert(tk.END, col[0].value)

import_button = tk.Button(headings_window, text='Import Headings', command=import_headings)
import_button.grid(row=4, column=0)

# Create a button for saving the changes and closing the window
def save_and_close():
    workbook.save(file_path)
    headings_window.destroy()

save_button = tk.Button(headings_window, text='Save', command=save_and_close)
save_button.grid(row=5, column=0)

# Run the Tkinter event loop
headings_window.mainloop()

print('Column headings saved successfully!')
